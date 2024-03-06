import configparser, os, time, glob, openpyxl
from functions.download_plan import download_file
from functions.unmerge_all import unmerge_all
from functions.unmerge_all import convert_to_xlsm
from functions.log import log
from dotenv import load_dotenv
import datetime
import base64
import pandas as pd
from openpyxl import load_workbook
import json, shutil, psycopg2, schedule
from functions.helpers import (
    get_datetime_range,
    format_time_range,
    calculate_checksum,
    db_insert_mongodb,
    get_latest_checksum,
)
load_dotenv()

def refresh_plan():
    start_time = time.time()
    username = os.getenv("PUW_USERNAME")
    password = base64.b64decode(os.getenv("PUW_PASSWORD")).decode("utf-8")
    try:
        os.remove("swiezy.xlsx")

    except:
        pass
    directory = os.getcwd()
    before_files = os.listdir(directory)
    # log("Trying to remove the old .xlsx file ")

    # Download the file
    try:
        download_file(username, password, directory)
        log("Successfully downloaded the file")
    except Exception as e:
        log(f"Error while downloading the file: {str(e)}")
        exit()

    # Move the new files to a folder with a timestamp
    newest_folder = directory
    xlsx_file = [file for file in os.listdir(newest_folder) if file.endswith(".xlsx")][
        0
    ]
    if not xlsx_file:
        log("No .xlsm files found in the directory.")
        exit()
    xlsx_file_new = "swiezy.xlsx"
    new_path = os.path.join(directory, xlsx_file_new)
    try:
        os.rename(os.path.join(directory, xlsx_file), new_path)
        log("Renamed the file")
    except Exception as e:
        log(f"Error while renaming the file: {str(e)}")
        exit()
    checksum_before = calculate_checksum(new_path)
    log(f"Checksum from downloaded file: {checksum_before}")
    checksum_db = get_latest_checksum()
    log(f"Latest checksum from database: {checksum_db}")
    if checksum_before != checksum_db:
        log("Changes detected")
        convert_to_xlsm("swiezy.xlsx")
        log("Unmerging all cells")
        new_path = os.path.join(directory, "swiezy.xlsm")
        unmerge_all(new_path)
        time.sleep(5)

        workbook = openpyxl.load_workbook(new_path)
        worksheet = workbook["INF st I"]
        row_number = 4
        current_date = datetime.datetime.now().strftime("D%d-%H-%M-%S")
        path_with_date = os.getcwd() + f"\{current_date}"
        os.mkdir(path_with_date)
        json_data = []
        log("Splitting the file into groups")
        for grupa in range(1, 7):
            log(f"Splitting the file for gr.{grupa}")

            column_letters = []
            # Iterate over the columns in the row
            for column in worksheet.iter_cols(min_row=row_number, max_row=row_number):
                # Get the column letter
                column_letter = column[0].column_letter

                # Iterate over the cells in the column
                for cell in column:
                    # Check if the cell contains the search term
                    if f"grupa {grupa}" in str(cell.value).lower():
                        # Add the cell value to the list of values
                        column_letters.append(column_letter)

            # Print the dictionary to the console
            print(f"Column letters for gr.{grupa} -  {column_letters}")
            df = pd.read_excel(new_path, sheet_name="INF st I", skiprows=4)

            def excel_columns():
                n = 1
                while True:
                    number = n
                    column_name = ""
                    while number > 0:
                        number, remainder = divmod(number - 1, 26)
                        column_name = chr(65 + remainder) + column_name
                    yield column_name
                    n += 1

            column_generator = excel_columns()
            df.columns = [next(column_generator) for _ in range(len(df.columns))]
            plan = df[column_letters]
            first_column = df[df.columns[0]]
            combined_df = pd.concat([first_column, plan], axis=1)
            combined_df = combined_df.drop(combined_df.index[-1])
            weekdays = ["Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek"]
            # Repeat or truncate the list of weekdays to match the number of columns
            weekdays = (
                weekdays * ((len(combined_df.columns) - 1) // len(weekdays))
            ) + weekdays[: (len(combined_df.columns) - 1) % len(weekdays)]

            # Add an empty string at the beginning of the list to skip the first column
            weekdays = ["godziny"] + weekdays

            # Add the weekdays as a new row at the top of the DataFrame
            combined_df.loc[-1] = weekdays
            combined_df.index = combined_df.index + 1
            combined_df = combined_df.sort_index()
            combined_df.to_excel(
                f"{current_date}\plan-edited-{grupa}.xlsx", index=False, header=False
            )
            df = pd.read_excel(
                f"{current_date}\plan-edited-{grupa}.xlsx",
                sheet_name="Sheet1",
            )
            df.fillna("", inplace=True)
            df["godziny"] = df["godziny"].apply(format_time_range)
            df.apply(
                lambda row: [
                    f"{row.iloc[0]} {cell_value}" if i > 0 else cell_value
                    for i, cell_value in enumerate(row)
                ],
                axis=1,
            )
            relevant_columns = ["Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek"]
            types_of_course = [
                "wykład",
                "ćwiczenia",
                "laboratorium",
                "lektorat",
                "warsztat",
            ]
            for column_name in relevant_columns:
                for i, cell_value in enumerate(df[column_name]):
                    row_numbers = df.loc[df[column_name] == cell_value].index
                    parts = cell_value.split("-")
                    lesson_start, lesson_end = get_datetime_range(
                        df["godziny"][row_numbers[0]]
                        .replace("<sup>", "")
                        .replace("</sup>", "")
                    )
                    course_name = parts[0].strip()
                    cell_value = cell_value.replace(course_name, "")
                    df.at[
                            i, column_name
                        ] = f"<div><b>{course_name}</b><br>{cell_value}</div>"
            data = df.to_dict("records")
            json_data.append(data)
            try:
                with open(f"{path_with_date}\plan-edited-{grupa}.json", "w") as f:
                    json.dump(data, f)
                log(
                    f"Sucesfully saved plan for group {grupa} in {current_date}\plan-edited-{grupa}.json"
                )
            except:
                log("Failed to saved the plan in json")

        now = datetime.datetime.now()
        end_time = time.time()
        future = now + datetime.timedelta(minutes=15)
        execution_time = end_time - start_time
        db_insert_mongodb(json_data, checksum_before, round(execution_time,2))
        log("Successfully inserted the document into the database")
        print("Current time: ", now)
        print("Next update in 15 minutes at: ", future)

    else:
        try:
            os.remove("swiezy.xlsx")
        except Exception as e:
            log(f"Error while removing the file or there's is no file")

        try:
            os.remove("swiezy.xlsm")
        except Exception as e:
            log(f"Error while removing the file or there's is no file")

        log("No changes detected")  # Next check in
        now = datetime.datetime.now()
        future = now + datetime.timedelta(minutes=15)
        print("Current time: ", now)
        print("Next update in 15 minutes at: ", future)
        # Close the connection


refresh_plan()
schedule.every(15).minutes.do(refresh_plan)

while True:
    schedule.run_pending()
    time.sleep(15)
