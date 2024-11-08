from LessonPlanDownloader import LessonPlanDownloader
import os, requests, time
from colorama import init, Fore, Style
from difflib import SequenceMatcher


init(autoreset=True)  # Initialize colorama
import pandas as pd
import openpyxl
import os, re
import pymongo
from datetime import datetime


class LessonPlan(LessonPlanDownloader):
    def __init__(self, username, password, mongo_uri, plan_config, directory=""):
        super().__init__(username, password, directory, plan_config["download_url"])
        self.plan_config = plan_config
        self.sheet_name = plan_config["sheet_name"]
        self.plans_directory = os.path.join(
            os.getenv("PLANS_DIRECTORY", "lesson_plans"),
            self.plan_config["name"].replace(" ", "_"),
        )
        os.makedirs(self.plans_directory, exist_ok=True)
        self.converted_lesson_plan = None
        self.groups = plan_config["groups"]
        self.group_columns = {}
        self.save_to_mongodb = os.getenv("SAVE_TO_MONGODB", "true").lower() == "true"
        self.save_to_file = os.getenv("SAVE_TO_FILE", "true").lower() == "true"
        self.plans_directory = os.getenv("PLANS_DIRECTORY", "lesson_plans")
        self.schedule_type = plan_config.get(
            "category", "st"
        )  # Default to standard schedule

        if self.save_to_mongodb:
            try:
                self.mongo_client = pymongo.MongoClient(mongo_uri)
                self.db = self.mongo_client["Lesson_dev"]
                print("Successfully connected to MongoDB")
            except pymongo.errors.ConnectionFailure as e:
                print(f"Could not connect to MongoDB: {e}")
            except Exception as e:
                print(f"An error occurred: {e}")
        if plan_config.get("groups") is None:
            self.groups = {"cały kierunek": "all"}
        else:
            self.groups = plan_config["groups"]

        self.group_columns = {}

    def get_schedule_headers(self, num_columns):
        """Return appropriate headers based on schedule type and actual number of columns"""
        base_headers = {
            "st": ["Godziny", "Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek"],
            "nst": ["Godziny", "Piątek", "Sobota", "Niedziela"],
            "nst-online": ["Godziny", "Sobota", "Niedziela"],
        }

        # Get base headers for schedule type
        headers = base_headers.get(self.schedule_type, base_headers["st"])

        # If we have more columns than headers, add numbered columns
        while len(headers) < num_columns:
            headers.append(f"Column_{len(headers)}")

        # If we have more headers than columns, trim the headers
        headers = headers[:num_columns]

        return headers

    def process_and_save_plan(self):
        """Process and save the lesson plan, returns checksum if plan was processed"""
        new_checksum = self.download_file()
        if not new_checksum:
            print("Failed to download file.")
            return None

        should_process = True

        if self.save_to_mongodb:
            # Use plan-specific collection
            collection_name = f"plans_{self.plan_config['name'].lower().replace(' ', '_').replace('-', '_')}"
            collection = self.db[collection_name]
            # Check MongoDB for changes
            latest_plan = collection.find_one(
                {"plan_name": self.plan_config["name"]}, 
                sort=[("timestamp", -1)]
            )

            latest_checksum = (
                latest_plan.get("checksum").split("_")[0]
                if latest_plan and latest_plan.get("checksum")
                else None
            )

            if latest_checksum and latest_checksum == new_checksum:
                print(f"Plan has not changed (MongoDB check in {collection_name}, checksum: {new_checksum}).")
                return False
            else:
                print(f"Plan has changed or no previous plan found (old checksum: {latest_checksum}, new checksum: {new_checksum})")
                should_process = True

        if should_process:
            print(f"Processing plan for {self.plan_config['name']}")
            
            try:
                # Always process the downloaded file
                self.unmerge_and_fill_data()
                self.clean_excel_file()

                # Process groups
                self.find_group_columns_with_similarity()
                
                # Get all groups from instance
                groups_to_process = self.groups.keys() if self.groups else []
                
                processed_groups = []
                failed_groups = []
                
                # Process each group
                for group_name in groups_to_process:
                    print(f"\nProcessing group: {group_name}")
                    try:
                        df_group = self.get_lessons_for_group(group_name)
                        if df_group is not None and not df_group.empty:
                            # Save group data
                            self.save_group_lessons(group_name, df_group)
                            processed_groups.append(group_name)
                            print(f"Successfully processed group: {group_name}")
                        else:
                            print(f"No data available for group: {group_name}")
                            failed_groups.append(group_name)
                    except Exception as group_error:
                        print(f"Error processing group {group_name}: {str(group_error)}")
                        failed_groups.append(group_name)
                        continue

                # Print summary
                print("\nProcessing Summary:")
                print(f"Successfully processed groups: {', '.join(processed_groups)}")
                if failed_groups:
                    print(f"Failed to process groups: {', '.join(failed_groups)}")

                # Save to MongoDB if enabled
                if self.save_to_mongodb and processed_groups:
                    self.convert_to_html_and_save_to_db(new_checksum)
                
                return new_checksum

            except Exception as e:
                print(f"Error processing plan: {str(e)}")
                import traceback
                traceback.print_exc()
                return None

        return new_checksum


    def get_converted_lesson_plan(self):
        return self.converted_lesson_plan

    def get_groups(self):
        return self.groups

    def unmerge_and_fill_data(self):
        if not self.file_save_path:
            print("No file has been downloaded yet. Please run download_file() first.")
            return False
        wb = openpyxl.load_workbook(self.file_save_path)
        ws = wb[self.sheet_name]

        merged_cells = list(ws.merged_cells)

        for merged_range in merged_cells:
            cell_range = str(merged_range)
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(cell_range)

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = merged_value

        dir_path = os.path.dirname(self.file_save_path)
        file_name = os.path.basename(self.file_save_path)
        new_file_name = "unmerged_" + file_name
        self.converted_lesson_plan = os.path.join(dir_path, new_file_name)

        wb.save(self.converted_lesson_plan)
        print(f"Unmerged file saved as: {self.converted_lesson_plan}{Style.RESET_ALL}")
        return True

    @staticmethod
    def clean_text(text):
        if isinstance(text, str):
            text = text.replace("\n", " ")
            text = re.sub(" +", " ", text)
            text = text.replace("\t", " ")
            return text.strip()
        return text

    def clean_excel_file(self):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return False

        file_name, file_extension = os.path.splitext(self.converted_lesson_plan)
        temp_file = file_name + "_temp" + file_extension

        try:
            with pd.ExcelFile(self.converted_lesson_plan) as xls:
                sheet_names = xls.sheet_names
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                for sheet_name in sheet_names:
                    # Read data into DataFrame
                    df = pd.read_excel(xls, sheet_name=sheet_name)

                    # Clean text data -
                    # df = df.apply(self.clean_text)

                    # Create new worksheet
                    ws = wb.create_sheet(title=sheet_name)

                    # Write headers
                    for col_num, value in enumerate(df.columns.values, 1):
                        ws.cell(row=1, column=col_num, value=value)

                    # Write data
                    for row_num, row in enumerate(df.values, 2):
                        for col_num, value in enumerate(row, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            # Reset all formatting
                            cell.font = openpyxl.styles.Font()
                            cell.fill = openpyxl.styles.PatternFill()
                            cell.border = openpyxl.styles.Border()
                            cell.alignment = openpyxl.styles.Alignment(wrap_text=False)
                            cell.number_format = "General"

                # Save workbook
                wb.save(temp_file)

            # Wait briefly for file operations to complete
            time.sleep(1)

            # Replace original file with cleaned version
            max_attempts = 5
            for attempt in range(max_attempts):
                try:
                    os.remove(self.converted_lesson_plan)
                    os.rename(temp_file, self.converted_lesson_plan)
                    break
                except PermissionError:
                    if attempt < max_attempts - 1:
                        time.sleep(1)
                    else:
                        raise

            print(f"Cleaned file saved as: {self.converted_lesson_plan}")
            return True

        except Exception as e:
            print(f"An error occurred while cleaning the file: {str(e)}")
            if os.path.exists(temp_file):
                os.remove(temp_file)
            return False

    def find_group_columns(self):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return False

        try:
            df = pd.read_excel(self.converted_lesson_plan, sheet_name=self.sheet_name)
            for key, value in self.groups.items():
                columns = df.columns[df.isin([value]).any()].tolist()
                self.group_columns[key] = columns

            return self.group_columns

        except Exception as e:
            print(f"An error occurred while finding group columns: {str(e)}")
            return None

    def find_group_columns_with_similarity(self):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return False

        def is_matching_group(text, pattern, similarity_threshold=1.0):
            """
            Sprawdza czy tekst odpowiada wzorcowi grupy z zadanym progiem podobieństwa
            """
            if not isinstance(text, str) or not isinstance(pattern, str):
                return False

            # Normalizacja tekstu
            def normalize_text(t):
                t = " ".join(t.split())
                return t.lower().strip()

            text_normalized = normalize_text(text)
            pattern_normalized = normalize_text(pattern)

            # 1. Najpierw sprawdź dokładne dopasowanie
            if text_normalized == pattern_normalized:
                print(
                    f"Exact match found after normalization for '{text}' and '{pattern}'"
                )
                return True

            # 2. Jeśli nie ma dokładnego dopasowania, sprawdź podobieństwo
            similarity = SequenceMatcher(
                None, text_normalized, pattern_normalized
            ).ratio()
            print(f"Comparing '{text}' with '{pattern}' - Similarity: {similarity}")
            return similarity >= similarity_threshold

        def verify_columns(columns, schedule_type):
            """
            Sprawdza czy znalezione kolumny są prawidłowe dla danego typu planu
            """
            expected_columns = {
                "st": 6,  # Godziny + 5 dni
                "nst": 4,  # Godziny + 3 dni
                "nst-online": 3,  # Godziny + 2 dni
            }

            # Sprawdź liczbę kolumn
            if len(columns) != expected_columns.get(schedule_type, 6):
                return False

            # Sprawdź czy nie ma kolumn numerowanych
            return not any("Column_" in col for col in columns)

        try:
            df = pd.read_excel(self.converted_lesson_plan, sheet_name=self.sheet_name)
            print("\nSearching for group columns...")
            print("Available columns:", df.columns.tolist())

            # Jeśli mamy grupę "cały kierunek", używamy poprzedniej logiki
            if len(self.groups) == 1 and "cały kierunek" in self.groups:
                print("Processing entire course without group division")
                days = {
                    "st": ["PONIEDZIAŁEK", "WTOREK", "ŚRODA", "CZWARTEK", "PIĄTEK"],
                    "nst": ["PIĄTEK", "SOBOTA", "NIEDZIELA"],
                    "nst-online": ["SOBOTA", "NIEDZIELA"],
                }
                schedule_days = days.get(self.schedule_type, days["st"])

                matching_columns = []
                used_columns = set()

                for column in df.columns:
                    if column in used_columns:
                        continue

                    unique_values = df[column].dropna().unique()
                    for value in unique_values:
                        value_str = str(value).upper().strip()
                        if value_str in schedule_days:
                            matching_columns.append(column)
                            used_columns.add(column)
                            break

                matching_columns.sort(
                    key=lambda x: int(x.split(".")[-1]) if "." in x else 0
                )
                self.group_columns["cały kierunek"] = matching_columns
                return self.group_columns

            # Standardowa logika dla zdefiniowanych grup
            used_columns = set()
            group_columns = {}
            backup_columns = {}

            for group_name, group_identifier in self.groups.items():
                print(
                    f"\nProcessing group: {group_name} (identifier: {group_identifier})"
                )

                # Szukamy od dokładnego dopasowania do minimalnego progu
                for similarity_threshold in [x / 100.0 for x in range(100, 84, -1)]:
                    matching_columns = []
                    used_columns = set()

                    for column in df.columns:
                        if column in used_columns:
                            continue

                        print(f"\nChecking column: {column}")
                        unique_values = df[column].dropna().unique()

                        for value in unique_values:
                            if is_matching_group(
                                str(value), group_identifier, similarity_threshold
                            ):
                                matching_columns.append(column)
                                used_columns.add(column)
                                print(f"Found matching column: {column}")
                                break

                    if matching_columns:
                        # Jeśli znaleźliśmy kolumny, sprawdź czy są prawidłowe
                        if verify_columns(matching_columns, self.schedule_type):
                            group_columns[group_name] = matching_columns
                            print(
                                f"Found valid columns for {group_name}: {matching_columns}"
                            )
                            break
                        else:
                            # Zachowaj jako backup jeśli jeszcze nie mamy
                            if group_name not in backup_columns:
                                backup_columns[group_name] = matching_columns
                                print(
                                    f"Saving backup columns for {group_name}: {matching_columns}"
                                )

                # Jeśli nie znaleźliśmo prawidłowych kolumn, użyj backup
                if group_name not in group_columns and group_name in backup_columns:
                    group_columns[group_name] = backup_columns[group_name]
                    print(
                        f"Using backup columns for {group_name}: {backup_columns[group_name]}"
                    )
                elif group_name not in group_columns:
                    print(f"No columns found for {group_name}")

                self.group_columns = group_columns
                return self.group_columns

        except Exception as e:
            print(f"An error occurred while finding group columns: {str(e)}")
            return None

    def get_lessons_for_group(self, group_name):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return None

        if not self.group_columns:
            self.find_group_columns_with_similarity()

        try:
            print(f"\nReading Excel file for group: {group_name}")
            df = pd.read_excel(
                self.converted_lesson_plan, sheet_name=self.sheet_name, header=None
            )

            if group_name not in self.group_columns:
                print(f"Group '{group_name}' not found.")
                return None

            # Extract column numbers from the column names
            group_col_indices = []
            for col_name in self.group_columns[group_name]:
                try:
                    if "." in col_name:
                        col_number = int(col_name.split(".")[-1])
                        group_col_indices.append(col_number)
                except ValueError as e:
                    print(
                        f"Warning: Could not parse column number from {col_name}: {e}"
                    )
                    continue

            if not group_col_indices:
                print(f"Could not extract column numbers for group {group_name}")
                return None

            print(f"Found column indices for {group_name}: {group_col_indices}")

            # Add time column (always first column) and sort indices
            columns_to_extract = [0] + sorted(group_col_indices)
            print(f"Columns to extract: {columns_to_extract}")

            # Extract columns
            df_filtered = df.iloc[:, columns_to_extract].copy()

            # Remove semester information rows (improved logic)
            df_filtered = df_filtered[
                ~df_filtered.apply(
                    lambda row: row.astype(str)
                    .str.contains("semestr|zjazd", case=False)
                    .any(),
                    axis=1,
                )
            ]

            # Find the first row with time information
            time_row_index = None
            for idx, row in df_filtered.iterrows():
                if isinstance(row[0], str) and any(
                    pattern in row[0].lower()
                    for pattern in [
                        "godz",
                        "godziny",
                        "7",
                        "8",
                        "9",
                        "10",
                        "11",
                        "12",
                        "13",
                        "14",
                        "15",
                        "16",
                        "17",
                        "18",
                        "19",
                        "20",
                    ]
                ):
                    time_row_index = idx
                    break

            if time_row_index is not None:
                # Remove all rows before the time row
                df_filtered = df_filtered.iloc[time_row_index:]
            else:
                print("Warning: Could not find time row")

            # Remove header rows that contain "godz" or "GODZ"
            df_filtered = df_filtered[
                ~df_filtered[0]
                .astype(str)
                .str.contains("godz\.|GODZ\.", case=False, regex=True)
            ]

            # Improved empty row removal
            # Convert all values to string and check if they're empty or whitespace
            df_filtered = df_filtered[
                df_filtered.apply(
                    lambda row: any(
                        str(cell).strip() and str(cell).strip().lower() != "nan"
                        for cell in row
                    ),
                    axis=1,
                )
            ]

            # Remove rows where all group columns (excluding time column) are NaN
            df_filtered = df_filtered.dropna(subset=df_filtered.columns[1:], how="all")

            # Get appropriate headers
            headers = self.get_schedule_headers(len(df_filtered.columns))
            df_filtered.columns = headers

            # Reset index
            df_filtered = df_filtered.reset_index(drop=True)

            # Final validation of the data
            if df_filtered.empty:
                print(f"Warning: No data found for group {group_name}")
                return None

            # Verify that we have all expected time slots
            expected_time_slots = [
                "725- 810",
                "815- 900",
                "905- 950",
                "1000-1045",
                "1050- 1135",
                "1145- 1230",
                "1235- 1320",
                "1330- 1415",
                "1420- 1505",
                "1515- 1600",
                "1605- 1650",
                "1700- 1745",
                "1750- 1835",
                "1845- 1930",
                "1935- 2020",
                "2030- 2115",
            ]

            missing_slots = []
            for slot in expected_time_slots:
                if not any(
                    df_filtered["Godziny"].astype(str).str.contains(slot, regex=False)
                ):
                    missing_slots.append(slot)

            if missing_slots:
                print(f"Warning: Missing time slots for {group_name}: {missing_slots}")
            if not df_filtered.empty:
                last_row_time = str(df_filtered.iloc[-1]["Godziny"]).strip()
                # Sprawdź czy ostatni wiersz nie zawiera godziny
                if not any(
                    time_pattern in last_row_time
                    for time_pattern in [
                        "725-",
                        "815-",
                        "905-",
                        "1000-",
                        "1050-",
                        "1145-",
                        "1235-",
                        "1330-",
                        "1420-",
                        "1515-",
                        "1605-",
                        "1700-",
                        "1750-",
                        "1845-",
                        "1935-",
                        "2030-",
                    ]
                ):
                    # Usuń ostatni wiersz
                    df_filtered = df_filtered.iloc[:-1]
                    print(f"Removed last row with non-time value: {last_row_time}")
            return df_filtered

        except Exception as e:
            print(
                f"An error occurred while getting lessons for group '{group_name}': {str(e)}"
            )
            import traceback

            traceback.print_exc()
            return None

    def save_group_lessons(self, group_name, df):
        if df is None or df.empty:
            print(f"No data to save for group '{group_name}'.")
            return

        # Create a directory for group files if it doesn't exist
        group_dir = os.path.join(
            os.path.dirname(self.converted_lesson_plan), "group_lessons"
        )
        os.makedirs(group_dir, exist_ok=True)

        # Create a file name for the group
        file_name = f"{group_name.replace(' ', '_')}_lessons.xlsx"
        file_path = os.path.join(group_dir, file_name)

        try:
            # Truncate sheet name to 31 characters
            sheet_name = group_name[:31]

            # Save the DataFrame to an Excel file
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)

            print(f"Lessons for group '{group_name}' saved to {file_path}")
            if len(group_name) > 31:
                print(f"Note: Sheet name was truncated to '{sheet_name}'")
        except Exception as e:
            print(
                f"An error occurred while saving lessons for group '{group_name}': {str(e)}"
            )

    @staticmethod
    def get_column_letter(column_number):
        """Convert a column number to a column letter (A, B, C, ..., Z, AA, AB, ...)."""
        dividend = column_number
        column_letter = ""
        while dividend > 0:
            dividend, remainder = divmod(dividend - 1, 26)
            column_letter = chr(65 + remainder) + column_letter
        return column_letter

    def convert_to_html_and_save_to_db(self, checksum):
        if not self.group_columns:
            print("No group columns found. Please run find_group_columns() first.")
            return

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        plans_data = {
            "timestamp": current_datetime,
            "checksum": checksum,  # Just store the raw checksum
            "plan_name": self.plan_config["name"],
            "category": self.schedule_type,
            "groups": {},
        }

        # Create plans directory if saving to files is enabled
        if self.save_to_file:
            plan_specific_directory = os.path.join(
                self.plans_directory, self.plan_config["name"].replace(" ", "_")
            )
            os.makedirs(plan_specific_directory, exist_ok=True)
            self.plans_directory = plan_specific_directory

        for group_name in self.groups.keys():
            df = self.get_lessons_for_group(group_name)
            if df is None or df.empty:
                print(f"No data available for group '{group_name}'.")
                continue

            html = self.generate_html_table(df)
            plans_data["groups"][group_name] = html

            # Save DataFrame to file if enabled
            if self.save_to_file:
                file_name = f"{current_datetime.replace(':', '-')}_{group_name}.pkl"
                file_path = os.path.join(self.plans_directory, file_name)
                df.to_pickle(file_path)
                print(f"Lesson plan for {group_name} saved to file: {file_path}")

        # Save to MongoDB if enabled
        if self.save_to_mongodb:

            collection_name = f"plans_{self.plan_config['name'].lower().replace(' ','_').replace('-', '_')}"
            collection = self.db[collection_name]

            # Check if this checksum already exists in this collection
            existing_plan = collection.find_one({"checksum": checksum})
            if existing_plan:
                print(
                    f"Plan with checksum {checksum} alreadyexists in collection {collection_name}. Skipping save."
                )
                return

            # Insert the new plan
            collection.insert_one(plans_data)
            print(
                f"Lesson plans for {self.plan_config['name']} timestamp {current_datetime} saved to MongoDB collection        {collection_name} with checksum {checksum}."
            )

    def generate_html_table(self, df):
        html = "<table border='1'>\n"

        # Add header row
        html += "<tr>\n"
        for col in df.columns:
            html += f"<th>{' '.join([f'<b>{word}</b>' for word in col.split()])}</th>\n"
        html += "</tr>\n"

        # Add data rows
        for _, row in df.iterrows():
            html += "<tr>\n"
            for i, cell in enumerate(row):
                formatted_cell = self.format_cell(
                    cell, is_header=False, is_time_column=(i == 0)
                )
                html += f"<td>{formatted_cell}</td>\n"
            html += "</tr>\n"

        html += "</table>"
        return html

    def format_cell(self, cell, is_header, is_time_column):
        if pd.isna(cell):
            return ""
        cell = str(cell)
        if is_time_column and cell != "godziny":
            parts = cell.replace(" ", "").split("-")
            if len(parts) == 2:
                start, end = parts
                formatted_start = self.format_time(start)
                formatted_end = self.format_time(end)
                return f"{formatted_start} - {formatted_end}"
        return cell

    def format_time(self, time):
        if len(time) == 3:
            return f"{time[0]}<sup>{time[1:]}</sup>"
        elif len(time) == 4:
            return f"{time[:2]}<sup>{time[2:]}</sup>"
        return time

    def full_action(self):
        """Execute the complete processing workflow"""
        try:
            checksum = self.process_and_save_plan()
            if checksum:
                print(f"Successfully completed processing with checksum: {checksum}")
            else:
                print("Processing completed but no new data was saved (no changes or errors occurred)")
        except Exception as e:
            print(f"Error during full action: {str(e)}")
            import traceback
            traceback.print_exc()
