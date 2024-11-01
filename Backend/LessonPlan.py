from LessonPlanDownloader import LessonPlanDownloader
import os, requests, time
import pandas as pd
import openpyxl
import os, re
import pymongo
from datetime import datetime

class LessonPlan(LessonPlanDownloader):
    def __init__(self, username, password, mongo_uri, directory="", sheet_name="INF st II"):
        super().__init__(username, password, directory)
        self.sheet_name = sheet_name
        self.converted_lesson_plan = None
        self.groups = {
            'Technologie Webowe i Internet rzeczy grupa 1':'''Sp.: Technologie Webowe i Internet rzeczy
grupa 1
podział wg nazwisk: A-K''',
            'Technologie Webowe i Internet rzeczy grupa 2':'''Technologie Webowe i Internet rzeczy
grupa 2
podział wg nazwisk: L-Z''',
            'Technologie mobilne':'Sp.: Technologie mobilne',
            'Grafika komputerowa i projektowanie gier':'Sp.: Grafika komputerowa i projektowanie gier',
            'Cyberbezpieczeństwo i informatyka śledcza grupa 1':'''Sp.: Cyberbezpieczeństwo i informatyka śledcza
grupa 1
podział wg nazwisk:A-K''',
            'Cyberbezpieczeństwo i informatyka śledcza grupa 2': '''Sp.: Cyberbezpieczeństwo i informatyka śledcza
grupa 2
podział wg nazwisk: L-Z''',
        }
        self.group_columns = {}
        
        try:
            self.mongo_client = pymongo.MongoClient(mongo_uri)
            self.db = self.mongo_client["Lesson"]
            print("Successfully connected to MongoDB")
        except pymongo.errors.ConnectionFailure as e:
            print(f"Could not connect to MongoDB: {e}")
        except Exception as e:
            print(f"An error occurred: {e}")
            
    def process_and_save_plan(self):
        new_checksum = self.download_file()
        print(f"New checksum: {new_checksum}")
        if not new_checksum:
            print("Failed to download file.")
            return
        
        latest_plan = self.db.plans.find_one(sort=[("timestamp", -1)])
        print("Old checksum: ", latest_plan.get("checksum") if latest_plan else "None")
        if latest_plan and latest_plan.get("checksum") == new_checksum:
            print("Plan has not changed. Skipping processing.")
            return

        self.download_file()
        self.unmerge_and_fill_data()
        self.clean_excel_file()
        self.find_group_columns()
        for group in self.groups.keys():
            df_group = self.get_lessons_for_group(group)
            if df_group is not None and not df_group.empty:
                self.save_group_lessons(group, df_group)
            else:
                print(f"No data available for group '{group}'.")
        self.convert_to_html_and_save_to_db(new_checksum)
        return True
        
    def get_converted_lesson_plan(self):
        return self.converted_lesson_plan

    def get_groups(self):
        return self.groups
        
    def unmerge_and_fill_data(self):
        if not self.file_save_path:
            print("No file has been downloaded yet. Please run download_file() first.")
            return False
        wb = openpyxl.load_workbook(self.file_save_path)
        ws = wb[self.sheet_name]
        
        merged_cells = list(ws.merged_cells)
    
        for merged_range in merged_cells:
            cell_range = str(merged_range)
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(cell_range)
    
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = merged_value
    
        dir_path = os.path.dirname(self.file_save_path)
        file_name = os.path.basename(self.file_save_path)
        new_file_name = 'unmerged_' + file_name
        self.converted_lesson_plan = os.path.join(dir_path, new_file_name)
        
        wb.save(self.converted_lesson_plan)
        print(f"Unmerged file saved as: {self.converted_lesson_plan}")
        return True

    @staticmethod
    def clean_text(text):
        if isinstance(text, str):
            text = text.replace('\n', ' ')
            text = re.sub(' +', ' ', text)
            text = text.replace('\t', ' ')
            return text.strip()
        return text

    def clean_excel_file(self):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return False

        file_name, file_extension = os.path.splitext(self.converted_lesson_plan)
        temp_file = file_name + '_temp' + file_extension

        try:
            with pd.ExcelFile(self.converted_lesson_plan) as xls:
                sheet_names = xls.sheet_names

                with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                    for sheet_name in sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        df = df.apply(self.clean_text)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

            time.sleep(1)
            
            max_attempts = 5
            for attempt in range(max_attempts):
                try:
                    os.remove(self.converted_lesson_plan)
                    os.rename(temp_file, self.converted_lesson_plan)
                    break
                except PermissionError:
                    if attempt < max_attempts - 1:
                        time.sleep(1)  
                    else:
                        raise 
            
            print(f"Cleaned file saved as: {self.converted_lesson_plan}")
            return True
        
        except Exception as e:
            print(f"An error occurred while cleaning the file: {str(e)}")
            if os.path.exists(temp_file):
                os.remove(temp_file)  
            return False

    def find_group_columns(self):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return False

        try:
            df = pd.read_excel(self.converted_lesson_plan, sheet_name=self.sheet_name)
            for key, value in self.groups.items():
                columns = df.columns[df.isin([value]).any()].tolist()
                self.group_columns[key] = columns

            print("Group columns found:")
            for group, columns in self.group_columns.items():
                print(f"{group}: {', '.join(columns)}")

            return self.group_columns

        except Exception as e:
            print(f"An error occurred while finding group columns: {str(e)}")
            return None

    def get_lessons_for_group(self, group_name):
        if not self.converted_lesson_plan:
            print("No converted file found. Please run unmerge_and_fill_data() first.")
            return None

        if not self.group_columns:
            self.find_group_columns()

        try:
            df = pd.read_excel(self.converted_lesson_plan, sheet_name=self.sheet_name, header=None)

            if group_name not in self.group_columns:
                print(f"Group '{group_name}' not found.")
                return None

            group_col_indices = [df.columns[df.iloc[0] == col].tolist()[0] for col in self.group_columns[group_name]]
            columns_to_extract = [0] + group_col_indices  

            df_filtered = df.iloc[:, columns_to_extract]

            df_filtered = df_filtered[~df_filtered.apply(lambda row: row.astype(str).str.contains('INFORMATYKA III SEMESTR rok akademicki 2024/2025').any(), axis=1)]

            df_filtered = df_filtered.iloc[4:-1]

            df_filtered = df_filtered.dropna(subset=df_filtered.columns[1:], how='all')

            headers = ['Godziny', 'Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek']
            df_filtered.columns = headers[:len(df_filtered.columns)]

            df_filtered = df_filtered.reset_index(drop=True)

            print(f"DataFrame for group '{group_name}' created successfully.")
            print(f"Columns: {', '.join(df_filtered.columns.astype(str))}")
            print(f"Shape: {df_filtered.shape}")

            return df_filtered

        except Exception as e:
            print(f"An error occurred while getting lessons for group '{group_name}': {str(e)}")
            return None

    def save_group_lessons(self, group_name, df):
        if df is None or df.empty:
            print(f"No data to save for group '{group_name}'.")
            return

        group_dir = os.path.join(os.path.dirname(self.converted_lesson_plan), "group_lessons")
        os.makedirs(group_dir, exist_ok=True)

        file_name = f"{group_name.replace(' ', '_')}_lessons.xlsx"
        file_path = os.path.join(group_dir, file_name)

        try:
            sheet_name = group_name[:31]

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            print(f"Lessons for group '{group_name}' saved to {file_path}")
            if len(group_name) > 31:
                print(f"Note: Sheet name was truncated to '{sheet_name}'")
        except Exception as e:
            print(f"An error occurred while saving lessons for group '{group_name}': {str(e)}")
    @staticmethod
    def get_column_letter(column_number):
        """Convert a column number to a column letter (A, B, C, ..., Z, AA, AB, ...)."""
        dividend = column_number
        column_letter = ''
        while dividend > 0:
            dividend, remainder = divmod(dividend - 1, 26)
            column_letter = chr(65 + remainder) + column_letter
        return column_letter

    def convert_to_html_and_save_to_db(self, checksum):
        if not self.group_columns:
            print("No group columns found. Please run find_group_columns() first.")
            return

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        plans_data = {
            "timestamp": current_datetime,
            "checksum": checksum,
            "groups": {}
        }

        for group_name in self.groups.keys():
            df = self.get_lessons_for_group(group_name)
            if df is None or df.empty:
                print(f"No data available for group '{group_name}'.")
                continue

            html = self.generate_html_table(df)
            plans_data["groups"][group_name] = html

        self.db.plans.insert_one(plans_data)
        print(f"Lesson plans for timestamp {current_datetime} saved to MongoDB with checksum {checksum}.")

    def generate_html_table(self, df):
        html = "<table border='1'>\n"
        
        html += "<tr>\n"
        for col in df.columns:
            html += f"<th>{' '.join([f'<b>{word}</b>' for word in col.split()])}</th>\n"
        html += "</tr>\n"

        for _, row in df.iterrows():
            html += "<tr>\n"
            for i, cell in enumerate(row):
                formatted_cell = self.format_cell(cell, is_header=False, is_time_column=(i==0))
                html += f"<td>{formatted_cell}</td>\n"
            html += "</tr>\n"

        html += "</table>"
        return html

    def format_cell(self, cell, is_header, is_time_column):
        if pd.isna(cell):
            return ""
        cell = str(cell)
        if is_time_column and cell != "godziny":
            parts = cell.replace(' ', '').split('-')
            if len(parts) == 2:
                start, end = parts
                formatted_start = self.format_time(start)
                formatted_end = self.format_time(end)
                return f"{formatted_start} - {formatted_end}"
        return cell

    def format_time(self, time):
        if len(time) == 3:
            return f"{time[0]}<sup>{time[1:]}</sup>"
        elif len(time) == 4:
            return f"{time[:2]}<sup>{time[2:]}</sup>"
        return time
    def full_action(self):
        self.process_and_save_plan()
        
